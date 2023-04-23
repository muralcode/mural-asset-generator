"""
Author/Engineer: Lerato Mokoena

This file houses the code class ClientData, and its purpose is to convert client input from Excel format into usable
data structures for white-labelling apps.
"""
import openpyxl

STRINGS_SHEET = 'text'
FAQS_SHEET = 'faqs'
COLORS_SHEET = 'colors'


class ClientData:
    """
    A class used to extract client-supplied values from an MS Excel document.

    Attributes
    ----------
    none

    Methods
    -------
    get_strings()
        getter that returns the client strings loaded from workbook in __init__
    get_faqs()
        getter that returns the client faqs loaded from workbook in __init__
    get_colors()
        getter that returns the client colors loaded from workbook in __init__
    get_name()
        getter that returns the name we use for the client
    """
    def __init__(self, path_to_file, client):
        workbook = openpyxl.load_workbook(path_to_file)
        self.strings = _load_strings(workbook, client)
        self.faqs = _load_faqs(workbook, client)
        self.colors = _load_colors(workbook, client)
        self.name = client

    def get_strings(self):
        """
        :return: localisation strings
        """
        return self.strings

    def get_faqs(self):
        """
        :return: client frequently asked questions
        """
        return self.faqs

    def get_colors(self):
        """
        :return: client color scheme
        """
        return self.colors

    def get_name(self):
        """
        :return: name of the client
        """
        return self.name


def _load_strings(workbook, client):
    """
    :param workbook: the workbook where the strings are located
    :param client: the client
    :return: the localised strings
    """
    if STRINGS_SHEET not in workbook.sheetnames:
        print('No sheet named "{0}" was found, strings files not created for "{1}"'.format(STRINGS_SHEET, client))
        return None

    sheet = workbook.get_sheet_by_name(STRINGS_SHEET)

    localised_strings = []
    for i in range(1, len(sheet.rows[0])):
        lang = sheet.rows[0][i].value

        strings_list = []
        for row in sheet.rows[1:]:
            name = row[0].value
            if name is not None:
                text = row[i].value.strip() if row[i].value is not None else ''
                strings_list.append((name, text))
        localised_strings.append((lang, strings_list, i == 1))
    return localised_strings


def _load_faqs(workbook, client):
    """
    :param workbook: the workbook where the faqs are located
    :param client: the client
    :return: the localised faqs
    """
    if FAQS_SHEET not in workbook.sheetnames:
        print('No sheet named "{0}" was found, faq files not created for "{1}"'.format(FAQS_SHEET, client))
        return None

    sheet = workbook.get_sheet_by_name(FAQS_SHEET)

    localised_faqs = []
    for i in range(1, len(sheet.rows[0])):
        lang = sheet.rows[0][i].value
        if lang is None:
            break

        rows = sheet.rows[1:]

        faqs_list = []
        for j in range(0, len(rows), 2):
            if rows[j][i].value is None:
                break

            text = rows[j + 1][i].value.replace('\n', '\\n')
            faqs_list.append((rows[j][i].value, text))

        localised_faqs.append((lang, faqs_list, i == 1))
    return localised_faqs


def _load_colors(workbook, client):
    """
    :param workbook: the workbook where the colors are located
    :param client: the client
    :return: the color schemes of the app
    """
    if COLORS_SHEET not in workbook.sheetnames:
        print('No sheet named "{0}" was found, colors files not created for "{1}"'.format(COLORS_SHEET, client))
        return None

    sheet = workbook.get_sheet_by_name(COLORS_SHEET)

    colors = {}
    for row in sheet.rows:
        name = row[0].value
        if name is not None:
            text = row[1].value if row[1].value is not None else ''
            colors[name] = text
    return colors
